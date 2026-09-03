"""NESDC 주별 GPP — 1인당 지역내총생산, 산업구성, 그리고 인구 분모.

DOPA는 SPA라 정적 다운로드가 불가능해, v1의 인구 분모는 여기서 가져온다.

2026-08-26 실측 확인 (Sprint 1)
  · 페이지가 **세션 쿠키 없이는 자기 자신으로 302 무한 리다이렉트**한다.
    requests.Session()으로 한 번 받아두면 이후 요청이 통과한다.
  · 파일 목록은 서버 렌더가 아니라 DataTable 초기화 스크립트 안의 JSON 리터럴이다
    (`jQuery('#dataTable').DataTable({ ... data: [ {...}, ... ] })`).
    각 항목의 download_file에 `?p=<post>&ddl=<file>` 링크가 들어 있고,
    그 링크가 wp-content/uploads의 실제 XLSX로 302한다.
  · 최신본: "Table of Gross Regional and Provincial Product 2024 (Excel)"
    → GPP-2024-On-Web-1995-2024.xlsx (2.2MB, 갱신 2026-03-31)
  · 시트 12개. 필요한 것은 둘.
      - `PER CAPITA`  : 77개 주 × (GPP 백만바트, 인구 천명, 1인당 GPP 바트)
      - 권역 시트 7개 : 주별 블록에 연도열 + 'Agriculture' / 'Gross provincial product (GPP)'
                        → 농림어업 비중을 여기서 만든다
  · 주 코드(0101 등)는 NESDC 내부 코드다. TIS-1099가 아니므로 이름으로 조인한다.

한계 (v1에서 해결하지 못함)
  · **도시화율(urbanization_rate)은 이 공표물에 없다.** 디지털 준비도 하향추정의
    가중치 키라서, 이것이 없으면 디지털 축이 통째로 결측이 된다. DATA_SOURCES.md 참조.
"""

from __future__ import annotations

import io
import json
import re
import time
from datetime import datetime

import openpyxl
import pandas as pd
import requests
from pathlib import Path

PAGE = "https://www.nesdc.go.th/en/info/gross-regional-and-provincial-product-gpp/"
BROWSER_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
)
HEADERS = {"User-Agent": BROWSER_UA, "Accept-Language": "en-US,en;q=0.9"}

# NESDC는 5xx를 간헐적으로 낸다 (2026-08-27 관찰: 502 Bad Gateway).
# **필수 소스**라 여기서 죽으면 월간 갱신 전체가 멈춘다 — 네트워크·5xx는 재시도한다.
# 4xx는 재시도하지 않는다: 경로가 틀린 것을 계속 두드려봐야 소용없다.
RETRIES = 3
BACKOFF_S = (3, 10, 25)

# 재시도로도 안 될 때를 위한 로컬 캐시.
#
# **이 소스에만 캐시 폴백을 허용하는 이유는 공표 주기가 1년이기 때문이다.**
# 지난달에 받아 둔 GPP 2024 워크북과 오늘 받는 GPP 2024 워크북은 같은 파일이다 —
# 캐시를 쓴다고 신선도가 손해나지 않는다. BOT은 월간이므로 같은 논리가 성립하지 않고,
# 그래서 BOT에는 캐시 폴백이 없다 (깨지면 그냥 멈춘다).
#
# 캐시는 data/raw/ 에 둔다 (gitignore). 공개 저장소에 정부 공표 XLSX를 재배포하지
# 않기 위해서다 — CI에서는 actions/cache가 실행 간에 이 디렉터리를 이어 준다.
CACHE_PREFIX = "nesdc-gpp"

TITLE_RE = re.compile(
    r"Table of Gross Regional and Provincial Product\s+(\d{4})\s*\(Excel\)", re.I
)
HREF_RE = re.compile(r'href="([^"]+)"')
PER_CAPITA_SHEET = "PER CAPITA"
REGION_SHEETS = ["NE", "NO", "SO", "EA", "WE", "CE", "BKK&VIC"]

CURRENT_PRICE_MARKER = "GROSS PROVINCIAL PRODUCT AT CURRENT MARKET PRICES"
PROVINCE_HEADER_RE = re.compile(r"^(\d{4})\s*[-–]\s*(.+)$")
PER_CAPITA_ROW_RE = re.compile(r"^(\d{4})\s+(.+)$")
AGRICULTURE_ROW = "agriculture"
GPP_ROW = "gross provincial product (gpp)"

THOUSAND = 1_000.0        # 인구: 천 명 → 명
MB = 1_000_000.0          # GPP: 백만 바트 → 바트


class SourceUnavailable(RuntimeError):
    """소스에 도달하지 못했다. 파싱 오류(ValueError)와 구분한다."""


def _year_of(label) -> int | None:
    """'2024p' · '2023r' · 2019 → 연도. 잠정·수정 표시는 떼어낸다."""
    m = re.match(r"^\s*(\d{4})\s*[a-zA-Z]?\s*$", str(label))
    return int(m.group(1)) if m else None


def _find_download(html: str) -> tuple[str, int]:
    """DataTable 초기화 JSON에서 가장 최신 연도의 GPP 엑셀 링크를 고른다."""
    best: tuple[int, str] | None = None
    for m in TITLE_RE.finditer(html):
        year = int(m.group(1))
        # 이 항목의 download_file 필드는 제목 바로 앞에 온다. 주변 창에서 href를 집는다.
        # JS 안에 박힌 JSON이라 슬래시와 따옴표가 이스케이프되어 있다. 창 단위로 되돌린다.
        window = html[max(0, m.start() - 600): m.end() + 600]
        window = window.replace("\\/", "/").replace('\\"', '"')
        hrefs = [h for h in HREF_RE.findall(window) if "ddl=" in h]
        if not hrefs:
            continue
        if best is None or year > best[0]:
            best = (year, hrefs[0])
    if best is None:
        raise ValueError(
            "NESDC 페이지에서 'Table of Gross Regional and Provincial Product YYYY (Excel)' "
            "항목을 찾지 못했다. DataTable JSON의 형식이나 공표물 제목이 바뀌었을 수 있다."
        )
    return best[1], best[0]


def _get(session: requests.Session, url: str, what: str, timeout: int, **kw):
    """5xx·네트워크 오류만 재시도한다. 4xx는 즉시 올린다."""
    last: Exception | None = None
    for attempt in range(RETRIES):
        try:
            r = session.get(url, timeout=timeout, **kw)
            if 400 <= r.status_code < 500:
                r.raise_for_status()
            if r.status_code >= 500:
                raise requests.HTTPError(f"{r.status_code} from NESDC", response=r)
            return r
        except requests.HTTPError as e:
            if e.response is not None and 400 <= e.response.status_code < 500:
                raise SourceUnavailable(f"{what}: {e}") from e
            last = e
        except requests.RequestException as e:
            last = e
        if attempt < RETRIES - 1:
            print(f"         WARN  {what} 재시도 {attempt + 1}/{RETRIES - 1}: {last}")
            time.sleep(BACKOFF_S[attempt])
    raise SourceUnavailable(f"{what}: {RETRIES}회 시도 후에도 실패했다 — {last}")


def _cached(raw_dir: Path | None) -> tuple[bytes, int, str] | None:
    """가장 최신 연도의 캐시본. 없으면 None."""
    if raw_dir is None or not raw_dir.exists():
        return None
    files = sorted(raw_dir.glob(f"{CACHE_PREFIX}-*.xlsx"))
    if not files:
        return None
    newest = files[-1]
    m = re.search(r"-(\d{4})\.xlsx$", newest.name)
    if not m:
        return None
    return newest.read_bytes(), int(m.group(1)), f"file://{newest}"


def _fetch_workbook(timeout: int = 180, raw_dir: Path | None = None) -> tuple[bytes, int, str, bool]:
    """(내용, 연도, 출처 URL, 캐시 사용 여부).

    라이브 실패 시 캐시로 내려간다. 캐시도 없으면 원래 예외를 그대로 올린다 —
    필수 소스이므로 그때는 빌드가 죽는 것이 맞다.
    """
    session = requests.Session()
    session.headers.update(HEADERS)

    try:
        page = _get(session, PAGE, "NESDC GPP 페이지에 도달하지 못했다", timeout)
        url, year = _find_download(page.text)
        xl = _get(session, url, f"NESDC 엑셀({url})을 받지 못했다", timeout,
                  headers={"Referer": PAGE})

        ctype = xl.headers.get("Content-Type", "")
        if "spreadsheet" not in ctype and not xl.content[:2] == b"PK":
            raise ValueError(f"NESDC 엑셀 응답이 스프레드시트가 아니다: Content-Type={ctype!r}")
    except SourceUnavailable as e:
        fallback = _cached(raw_dir)
        if fallback is None:
            raise
        content, year, src = fallback
        print(f"         WARN  NESDC 라이브 실패 → 캐시본을 쓴다 ({Path(src).name}).\n"
              f"               연 1회 공표물이라 내용은 같다. 사유: {e}")
        return content, year, src, True

    # 성공했으면 캐시를 갱신한다
    if raw_dir is not None:
        raw_dir.mkdir(parents=True, exist_ok=True)
        (raw_dir / f"{CACHE_PREFIX}-{year}.xlsx").write_bytes(xl.content)

    return xl.content, year, xl.url, False


def _parse_per_capita(wb) -> pd.DataFrame:
    """PER CAPITA 시트 → 주별 인구·GPP·1인당 GPP.

    시트에는 권역 합계 블록이 먼저 오고, 그 다음 주별 표가 온다.
    주 행만 'NNNN NAME' 형식이라 그것으로 가른다 (권역 행은 코드가 없다).
    """
    ws = wb[PER_CAPITA_SHEET]
    rows = []
    for r in ws.iter_rows(values_only=True):
        if len(r) < 5 or not isinstance(r[0], (int, float)):
            continue
        m = PER_CAPITA_ROW_RE.match(str(r[1] or "").strip())
        if not m:
            continue
        rows.append({
            "nesdc_code": m.group(1),
            "nesdc_name": m.group(2).strip(),
            "gpp_total": None if r[2] is None else float(r[2]) * MB,
            "population": None if r[3] is None else float(r[3]) * THOUSAND,
            "gpp_per_capita": None if r[4] is None else float(r[4]),
        })

    df = pd.DataFrame(rows)
    if len(df) != 77:
        raise ValueError(
            f"'{PER_CAPITA_SHEET}' 시트에서 주를 {len(df)}개 읽었다. 77개여야 한다. "
            f"시트 레이아웃이 바뀌었을 가능성이 높다."
        )
    if df["nesdc_code"].duplicated().any():
        raise ValueError("PER CAPITA 시트에 중복된 주 코드가 있다.")
    return df


def _parse_agriculture(wb) -> pd.DataFrame:
    """권역 시트들 → 최신 연도의 농림어업 비중(%).

    각 주가 두 블록을 갖는다: 경상가격(CURRENT MARKET PRICES)과 실질(CHAIN VOLUME).
    비중은 경상가격 블록에서만 읽는다.
    """
    rows = []
    for name in REGION_SHEETS:
        if name not in wb.sheetnames:
            raise ValueError(f"권역 시트 '{name}'이 없다. 워크북 구성이 바뀌었다: {wb.sheetnames}")
        grid = list(wb[name].iter_rows(values_only=True))

        starts = [
            i for i, r in enumerate(grid)
            if isinstance(r[0], str) and r[0].strip().upper().startswith(CURRENT_PRICE_MARKER)
        ]
        for i in starts:
            block = grid[i: i + 40]
            header = next(
                (r for r in block if isinstance(r[0], str) and PROVINCE_HEADER_RE.match(r[0].strip())),
                None,
            )
            if header is None:
                continue
            code, prov_name = PROVINCE_HEADER_RE.match(header[0].strip()).groups()

            years = next((r for r in block if any(_year_of(c) for c in r[1:])), None)
            if years is None:
                raise ValueError(f"{name}/{code}: 연도 헤더 행을 찾지 못했다.")
            col = max(
                (j for j in range(1, len(years)) if _year_of(years[j])),
                key=lambda j: _year_of(years[j]),
            )
            year = _year_of(years[col])

            def value(label: str):
                for r in block:
                    if isinstance(r[0], str) and r[0].strip().lower() == label:
                        v = r[col] if col < len(r) else None
                        return float(v) if isinstance(v, (int, float)) else None
                return None

            agri, gpp = value(AGRICULTURE_ROW), value(GPP_ROW)
            rows.append({
                "nesdc_code": code,
                "nesdc_name": prov_name.strip(),
                "agri_year": year,
                "gpp_agriculture_share": (
                    None if not agri or not gpp else round(agri / gpp * 100.0, 4)
                ),
            })

    df = pd.DataFrame(rows).drop_duplicates(subset="nesdc_code")
    if len(df) != 77:
        raise ValueError(f"권역 시트에서 주를 {len(df)}개 읽었다. 77개여야 한다.")
    return df


def load(config: dict) -> dict:
    alias_map: dict[str, int] = config["_alias_map"]
    from sources.admin_ref import norm      # 같은 정규화 규칙을 공유한다

    raw_dir = Path(config["_raw_dir"]) if config.get("_raw_dir") else None
    content, edition, url, from_cache = _fetch_workbook(raw_dir=raw_dir)
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    agri = _parse_agriculture(wb).drop(columns=["nesdc_name"])
    df = _parse_per_capita(wb).merge(agri, on="nesdc_code", how="left")
    df["tis1099_code"] = df["nesdc_name"].map(lambda n: alias_map.get(norm(n)))

    unmatched = df[df["tis1099_code"].isna()]
    if len(unmatched):
        names = ", ".join(unmatched["nesdc_name"])
        raise ValueError(
            f"NESDC 주명 {len(unmatched)}개가 크로스워크에 붙지 않는다: {names}. "
            f"data/reference/province_name_aliases.csv에 별칭을 추가할 것 "
            f"(이름을 코드에서 치환하지 마라)."
        )
    df["tis1099_code"] = df["tis1099_code"].astype(int)

    out = df[[
        "tis1099_code", "population", "gpp_per_capita", "gpp_total",
        "gpp_agriculture_share", "nesdc_name",
    ]]

    return {
        "provinces": out,
        "as_of": str(edition),
        "grade": "B",
        "source_url": PAGE,
        "file_url": url,
        "edition": edition,
        "from_cache": from_cache,
        "retrieved_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
    }


if __name__ == "__main__":
    import sys
    from pathlib import Path

    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    import yaml

    from sources import admin_ref

    cfg = yaml.safe_load(open(Path(__file__).resolve().parents[1] / "config.yaml", encoding="utf-8"))
    cfg["_reference_dir"] = str(Path(__file__).resolve().parents[2] / "data" / "reference")
    cfg["_alias_map"] = admin_ref.load(cfg)["alias_map"]
    r = load(cfg)
    print(json.dumps({k: v for k, v in r.items() if k != "provinces"}, ensure_ascii=False, indent=2))
    print(r["provinces"].head(10).to_string())
